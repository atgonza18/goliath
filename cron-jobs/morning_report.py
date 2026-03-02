#!/usr/bin/env python3
"""
Morning Report Delivery — standalone version.
Can be run from cron or manually; the bot scheduler uses its own copy.

Generates three report files (PDF, Excel, Markdown) and sends them as
Telegram document attachments with a SHORT notification message.

Files are saved to /opt/goliath/reports/YYYY-MM-DD-morning-report.{pdf,xlsx,md}
"""

import os
import re
import sqlite3
import sys
import requests
from datetime import datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo
from dotenv import load_dotenv

CT = ZoneInfo("America/Chicago")

REPO_ROOT = Path(__file__).resolve().parent.parent
CRON_REPORTS_DIR = REPO_ROOT / "cron-jobs" / "reports"
REPORTS_DIR = REPO_ROOT / "reports"
PROJECTS_DIR = REPO_ROOT / "projects"
MEMORY_DB_PATH = REPO_ROOT / "telegram-bot" / "data" / "memory.db"
FOLLOWUP_DB_PATH = REPO_ROOT / "telegram-bot" / "data" / "followup.db"

load_dotenv(REPO_ROOT / ".env")

TELEGRAM_BOT_TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN", "")
TELEGRAM_CHAT_ID = os.environ.get("REPORT_CHAT_ID", "") or os.environ.get("ALLOWED_CHAT_IDS", "")

# Project registry (mirrors bot/config.py)
PROJECTS = {
    "union-ridge":   "Union Ridge",
    "duff":          "Duff",
    "salt-branch":   "Salt Branch",
    "blackford":     "Blackford",
    "delta-bobcat":  "Delta Bobcat",
    "tehuacana":     "Tehuacana",
    "three-rivers":  "Three Rivers",
    "scioto-ridge":  "Scioto Ridge",
    "mayes":         "Mayes",
    "graceland":     "Graceland",
    "pecan-prairie": "Pecan Prairie",
    "duffy-bess":    "Duffy BESS",
}


# ------------------------------------------------------------------
# Data gathering (structured — for PDF / Excel / Markdown)
# ------------------------------------------------------------------

def gather_open_action_items() -> list[dict]:
    """Query the memory DB for all open (unresolved) action items."""
    if not MEMORY_DB_PATH.exists():
        return []
    try:
        conn = sqlite3.connect(str(MEMORY_DB_PATH))
        conn.row_factory = sqlite3.Row
        cursor = conn.execute(
            "SELECT id, created_at, project_key, summary, detail "
            "FROM memories WHERE category = 'action_item' AND resolved = 0 "
            "ORDER BY created_at DESC"
        )
        rows = [dict(row) for row in cursor.fetchall()]
        conn.close()
        return rows
    except Exception as e:
        print(f"Warning: could not query memory DB: {e}")
        return []


def gather_project_health() -> list[dict]:
    """Build project health data as a list of dicts."""
    results = []

    for slug, name in PROJECTS.items():
        project_path = PROJECTS_DIR / slug
        row = {
            "slug": slug,
            "name": name,
            "pod": 0,
            "schedule": 0,
            "constraints": 0,
            "constraints_open": 0,
            "schedule_status": "On Track",
            "key_risk": "None identified",
        }

        if project_path.exists():
            for folder in ["pod", "schedule", "constraints"]:
                folder_path = project_path / folder
                if folder_path.exists():
                    files = [
                        f for f in folder_path.rglob("*")
                        if f.is_file() and f.name != ".gitkeep"
                    ]
                    row[folder] = len(files)

            open_count = row["constraints_open"] or row["constraints"]
            if open_count > 10:
                row["schedule_status"] = "At Risk"
                row["key_risk"] = f"{open_count} open constraints"
            elif open_count > 5:
                row["schedule_status"] = "Monitor"
                row["key_risk"] = f"{open_count} open constraints"

        results.append(row)

    return results


def gather_constraint_movement() -> dict:
    """Get constraint movement data.

    NOTE: Constraint snapshot functionality has been removed. Constraint data
    is now accessed live via the Convex API / MCP server. Returns an empty
    structure for backward compatibility with report formatting.
    """
    return {
        "new": [], "resolved": [], "status_changed": [], "priority_changed": [],
        "total_current": 0,
        "per_project": {},
    }


def gather_followup_items() -> dict:
    """Query the follow-up DB for overdue and due-today items."""
    result = {"overdue": [], "due_today": []}
    if not FOLLOWUP_DB_PATH.exists():
        return result
    try:
        conn = sqlite3.connect(str(FOLLOWUP_DB_PATH))
        conn.row_factory = sqlite3.Row
        today = datetime.now(CT).strftime("%Y-%m-%d")

        cursor = conn.execute(
            "SELECT * FROM follow_ups WHERE follow_up_date < ? AND status = 'pending' "
            "ORDER BY follow_up_date ASC", (today,),
        )
        result["overdue"] = [dict(row) for row in cursor.fetchall()]

        cursor = conn.execute(
            "SELECT * FROM follow_ups WHERE follow_up_date = ? AND status = 'pending' "
            "ORDER BY created_at ASC", (today,),
        )
        result["due_today"] = [dict(row) for row in cursor.fetchall()]
        conn.close()
    except Exception as e:
        print(f"Warning: could not query follow-up DB: {e}")
    return result


def get_latest_scan_content() -> str | None:
    """Find the most recent daily scan report and return its content."""
    if not CRON_REPORTS_DIR.exists():
        return None
    reports = sorted(CRON_REPORTS_DIR.glob("*_daily_scan.md"), reverse=True)
    if not reports:
        return None
    try:
        content = reports[0].read_text(errors="replace")
        return content if content.strip() else None
    except Exception:
        return None


# ------------------------------------------------------------------
# PDF generation (ReportLab)
# ------------------------------------------------------------------

def generate_morning_pdf(
    output_path: Path, date_str: str, action_items: list[dict],
    project_health: list[dict], constraint_movement: dict,
    followup_items: dict, scan_content: str | None,
) -> bool:
    """Generate a professional morning report PDF. Returns True on success."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_LEFT, TA_CENTER
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import inch
        from reportlab.platypus import (
            BaseDocTemplate, Frame, HRFlowable,
            NextPageTemplate, PageBreak, PageTemplate,
            Paragraph, Spacer, Table, TableStyle,
        )
    except ImportError:
        print("Error: reportlab not installed")
        return False

    DSC_BLUE = colors.HexColor("#003366")
    ACCENT_BLUE = colors.HexColor("#336699")
    RED = colors.HexColor("#CC0000")
    AMBER = colors.HexColor("#CC8800")
    GREEN = colors.HexColor("#228B22")
    WHITE = colors.white
    BLACK = colors.black
    MID_GREY = colors.HexColor("#E0E0E0")
    DARK_GREY = colors.HexColor("#666666")
    TABLE_HEADER_BG = DSC_BLUE
    TABLE_ALT_ROW = colors.HexColor("#F2F6FA")

    _base = getSampleStyleSheet()

    def _s(name, **kw):
        return ParagraphStyle(name, parent=_base["Normal"], **kw)

    S_H1 = _s("H1", fontName="Helvetica-Bold", fontSize=14, textColor=DSC_BLUE,
               leading=20, spaceBefore=14, spaceAfter=6)
    S_H2 = _s("H2", fontName="Helvetica-Bold", fontSize=11, textColor=DSC_BLUE,
               leading=16, spaceBefore=10, spaceAfter=4)
    S_BODY = _s("Body", fontName="Helvetica", fontSize=9, textColor=BLACK,
                leading=12, spaceAfter=3)
    S_BODY_BOLD = _s("BodyBold", fontName="Helvetica-Bold", fontSize=9,
                     textColor=BLACK, leading=12, spaceAfter=3)
    S_SMALL = _s("Small", fontName="Helvetica", fontSize=7.5,
                 textColor=DARK_GREY, leading=10)
    S_TH = _s("TH", fontName="Helvetica-Bold", fontSize=8, textColor=WHITE, leading=10)
    S_TC = _s("TC", fontName="Helvetica", fontSize=8, textColor=BLACK, leading=10)
    S_ACTION = _s("Action", fontName="Helvetica", fontSize=8.5, textColor=BLACK,
                  leading=11, leftIndent=8, spaceAfter=2)

    w, h = letter
    now = datetime.now(CT)
    formatted_date = now.strftime("%B %d, %Y")

    def _esc(text):
        if not text:
            return ""
        return str(text).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

    def _first_page(canvas, doc):
        canvas.saveState()
        header_h = 90
        canvas.setFillColor(DSC_BLUE)
        canvas.rect(0, h - header_h, w, header_h, stroke=0, fill=1)
        canvas.setFillColor(ACCENT_BLUE)
        canvas.rect(0, h - header_h, w, 3, stroke=0, fill=1)
        canvas.setFillColor(WHITE)
        canvas.setFont("Helvetica-Bold", 22)
        canvas.drawString(0.75 * inch, h - 40, "GOLIATH Morning Report")
        canvas.setFont("Helvetica", 11)
        canvas.setFillColor(colors.HexColor("#B0C4DE"))
        canvas.drawString(0.75 * inch, h - 58, formatted_date)
        n_items = len(action_items)
        n_changes = len(constraint_movement.get("new", [])) + len(constraint_movement.get("resolved", []))
        canvas.setFont("Helvetica", 9)
        canvas.setFillColor(colors.HexColor("#8FAFD0"))
        canvas.drawString(0.75 * inch, h - 74,
                          f"{n_items} action items  |  {len(project_health)} projects  |  "
                          f"{n_changes} constraint changes (24h)")
        canvas.setStrokeColor(MID_GREY)
        canvas.setLineWidth(0.5)
        canvas.line(0.75 * inch, 0.5 * inch, w - 0.75 * inch, 0.5 * inch)
        canvas.setFillColor(DARK_GREY)
        canvas.setFont("Helvetica", 6.5)
        canvas.drawString(0.75 * inch, 0.35 * inch, "Generated by GOLIATH")
        canvas.drawRightString(w - 0.75 * inch, 0.35 * inch, f"Page {doc.page}")
        canvas.restoreState()

    def _later_pages(canvas, doc):
        canvas.saveState()
        canvas.setFillColor(DSC_BLUE)
        canvas.rect(0, h - 45, w, 45, stroke=0, fill=1)
        canvas.setFillColor(WHITE)
        canvas.setFont("Helvetica-Bold", 10)
        canvas.drawString(0.75 * inch, h - 30, "GOLIATH Morning Report")
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.HexColor("#B0C4DE"))
        canvas.drawRightString(w - 0.75 * inch, h - 30, formatted_date)
        canvas.setStrokeColor(MID_GREY)
        canvas.setLineWidth(0.5)
        canvas.line(0.75 * inch, 0.5 * inch, w - 0.75 * inch, 0.5 * inch)
        canvas.setFillColor(DARK_GREY)
        canvas.setFont("Helvetica", 6.5)
        canvas.drawString(0.75 * inch, 0.35 * inch, "Generated by GOLIATH")
        canvas.drawRightString(w - 0.75 * inch, 0.35 * inch, f"Page {doc.page}")
        canvas.restoreState()

    first_frame = Frame(0.75 * inch, 0.65 * inch, w - 1.5 * inch, h - 1.65 * inch, id="first")
    later_frame = Frame(0.75 * inch, 0.65 * inch, w - 1.5 * inch, h - 1.25 * inch, id="later")
    first_tmpl = PageTemplate(id="First", frames=[first_frame], onPage=_first_page)
    later_tmpl = PageTemplate(id="Later", frames=[later_frame], onPage=_later_pages)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    doc = BaseDocTemplate(str(output_path), pagesize=letter,
                          title=f"GOLIATH Morning Report - {formatted_date}",
                          author="GOLIATH System")
    doc.addPageTemplates([first_tmpl, later_tmpl])

    elements = []

    # ---- Open Action Items ----
    elements.append(Paragraph(f'OPEN ACTION ITEMS ({len(action_items)} pending)', S_H1))
    elements.append(HRFlowable(width="100%", thickness=0.5, color=MID_GREY,
                               spaceBefore=2, spaceAfter=6))
    if action_items:
        by_project = {}
        for item in action_items:
            key = item.get("project_key") or "general"
            by_project.setdefault(key, []).append(item)

        general = by_project.pop("general", [])
        if general:
            elements.append(Paragraph("General / System", S_H2))
            for item in general:
                date = item["created_at"][:10]
                elements.append(Paragraph(f'<b>[{date}]</b> {_esc(item.get("summary", ""))}', S_ACTION))

        for proj_key in sorted(by_project.keys()):
            proj_name = PROJECTS.get(proj_key, proj_key)
            elements.append(Paragraph(_esc(proj_name), S_H2))
            for item in by_project[proj_key]:
                date = item["created_at"][:10]
                elements.append(Paragraph(f'<b>[{date}]</b> {_esc(item.get("summary", ""))}', S_ACTION))
    else:
        elements.append(Paragraph('<i>No open action items.</i>', S_BODY))
    elements.append(Spacer(1, 8))

    # ---- Project Health Summary ----
    elements.append(Paragraph(f'PORTFOLIO HEALTH SUMMARY ({len(project_health)} projects)', S_H1))
    elements.append(HRFlowable(width="100%", thickness=0.5, color=MID_GREY,
                               spaceBefore=2, spaceAfter=6))

    health_header = [
        Paragraph("<b>Project</b>", S_TH), Paragraph("<b>POD</b>", S_TH),
        Paragraph("<b>Schedule</b>", S_TH), Paragraph("<b>Constraints</b>", S_TH),
        Paragraph("<b>Open</b>", S_TH), Paragraph("<b>Status</b>", S_TH),
        Paragraph("<b>Key Risk</b>", S_TH),
    ]
    health_data = [health_header]
    for p in project_health:
        status = p.get("schedule_status", "On Track")
        sc = "#CC0000" if status == "At Risk" else "#CC8800" if status == "Monitor" else "#228B22"
        health_data.append([
            Paragraph(_esc(p["name"]), S_TC),
            Paragraph(str(p.get("pod", 0)), S_TC),
            Paragraph(str(p.get("schedule", 0)), S_TC),
            Paragraph(str(p.get("constraints", 0)), S_TC),
            Paragraph(str(p.get("constraints_open", 0)), S_TC),
            Paragraph(f'<font color="{sc}"><b>{_esc(status)}</b></font>', S_TC),
            Paragraph(_esc(p.get("key_risk", "None")[:60]), S_TC),
        ])

    col_widths = [1.3 * inch, 0.5 * inch, 0.65 * inch, 0.75 * inch,
                  0.5 * inch, 0.7 * inch, 2.1 * inch]
    health_table = Table(health_data, colWidths=col_widths, repeatRows=1)
    health_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), TABLE_HEADER_BG),
        ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
        ("GRID", (0, 0), (-1, -1), 0.5, MID_GREY),
        ("ALIGN", (1, 0), (4, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, TABLE_ALT_ROW]),
    ]))
    elements.append(health_table)
    elements.append(Spacer(1, 10))
    elements.append(NextPageTemplate("Later"))

    # ---- Constraint Movement ----
    elements.append(Paragraph("CONSTRAINT MOVEMENT (24h)", S_H1))
    elements.append(HRFlowable(width="100%", thickness=0.5, color=MID_GREY,
                               spaceBefore=2, spaceAfter=6))
    total_c = constraint_movement.get("total_current", 0)
    elements.append(Paragraph(f'<i>{total_c} constraints tracked</i>', S_SMALL))

    new_items = constraint_movement.get("new", [])
    resolved_items = constraint_movement.get("resolved", [])
    total_changes = (len(new_items) + len(resolved_items)
                     + len(constraint_movement.get("status_changed", []))
                     + len(constraint_movement.get("priority_changed", [])))
    if total_changes == 0:
        elements.append(Paragraph('<i>No changes in the last 24 hours.</i>', S_BODY))
    else:
        elements.append(Paragraph(f'<b>{total_changes} change(s) detected</b>', S_BODY_BOLD))
        if new_items:
            elements.append(Paragraph(f'<b>New ({len(new_items)})</b>', S_H2))
            for item in new_items[:10]:
                elements.append(Paragraph(
                    f'<b>[{_esc(item.get("priority", "?"))}]</b> '
                    f'{_esc(item.get("project", "?"))}: {_esc(item.get("description", "?"))}', S_ACTION))
        if resolved_items:
            elements.append(Paragraph(f'<b>Resolved ({len(resolved_items)})</b>', S_H2))
            for item in resolved_items[:10]:
                elements.append(Paragraph(
                    f'{_esc(item.get("project", "?"))}: {_esc(item.get("description", "?"))}', S_ACTION))
    elements.append(Spacer(1, 10))

    # ---- Follow-Up Queue ----
    elements.append(Paragraph("FOLLOW-UP QUEUE", S_H1))
    elements.append(HRFlowable(width="100%", thickness=0.5, color=MID_GREY,
                               spaceBefore=2, spaceAfter=6))
    overdue = followup_items.get("overdue", [])
    due_today = followup_items.get("due_today", [])
    if not overdue and not due_today:
        elements.append(Paragraph('<i>No follow-ups due or overdue today.</i>', S_BODY))
    else:
        if overdue:
            elements.append(Paragraph(f'<font color="#CC0000"><b>Overdue ({len(overdue)})</b></font>', S_H2))
            for item in overdue[:12]:
                elements.append(Paragraph(
                    f'<b>[{_esc(str(item.get("project_key", "?")))}]</b> '
                    f'{_esc(str(item.get("commitment", ""))[:120])}', S_ACTION))
        if due_today:
            elements.append(Paragraph(f'<font color="#CC8800"><b>Due Today ({len(due_today)})</b></font>', S_H2))
            for item in due_today[:12]:
                elements.append(Paragraph(
                    f'<b>[{_esc(str(item.get("project_key", "?")))}]</b> '
                    f'{_esc(str(item.get("commitment", ""))[:120])}', S_ACTION))
    elements.append(Spacer(1, 10))

    # ---- Scan Findings ----
    if scan_content:
        elements.append(Paragraph("LATEST DAILY SCAN", S_H1))
        elements.append(HRFlowable(width="100%", thickness=0.5, color=MID_GREY,
                                   spaceBefore=2, spaceAfter=6))
        cleaned = re.sub(r"^#{1,4}\s+", "", scan_content, flags=re.MULTILINE)
        cleaned = cleaned.replace("**", "").replace("__", "")
        if len(cleaned) > 4000:
            cleaned = cleaned[:3800] + "\n\n... (truncated)"
        for line in cleaned.split("\n"):
            line = line.strip()
            if not line:
                elements.append(Spacer(1, 3))
            else:
                elements.append(Paragraph(_esc(line), S_BODY))

    try:
        doc.build(elements)
        print(f"PDF generated: {output_path}")
        return True
    except Exception as e:
        print(f"PDF generation failed: {e}")
        return False


# ------------------------------------------------------------------
# Excel generation (openpyxl)
# ------------------------------------------------------------------

def generate_morning_excel(
    output_path: Path, date_str: str, action_items: list[dict],
    project_health: list[dict], constraint_movement: dict,
) -> bool:
    """Generate morning report Excel workbook. Returns True on success."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("Error: openpyxl not installed")
        return False

    wb = Workbook()
    HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    BODY_FONT = Font(name="Calibri", size=10)
    ALT_ROW_FILL = PatternFill(start_color="F2F6FA", end_color="F2F6FA", fill_type="solid")
    RED_FONT = Font(name="Calibri", bold=True, color="CC0000", size=10)
    AMBER_FONT = Font(name="Calibri", bold=True, color="CC8800", size=10)
    GREEN_FONT = Font(name="Calibri", bold=True, color="228B22", size=10)
    THIN_BORDER = Border(
        left=Side(style="thin", color="D0D0D0"), right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"), bottom=Side(style="thin", color="D0D0D0"),
    )

    def _style_header(ws, row_num, col_count):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER

    def _style_body(ws, start_row, end_row, col_count):
        for row in range(start_row, end_row + 1):
            for col in range(1, col_count + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = BODY_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if (row - start_row) % 2 == 1:
                    cell.fill = ALT_ROW_FILL

    # Sheet 1: Action Items
    ws1 = wb.active
    ws1.title = "Action Items"
    headers1 = ["ID", "Date", "Description", "Status", "Project"]
    for col_idx, hdr in enumerate(headers1, 1):
        ws1.cell(row=1, column=col_idx, value=hdr)
    _style_header(ws1, 1, len(headers1))

    for row_idx, item in enumerate(action_items, 2):
        proj_key = item.get("project_key") or "general"
        proj_name = PROJECTS.get(proj_key, proj_key)
        ws1.cell(row=row_idx, column=1, value=item.get("id", ""))
        ws1.cell(row=row_idx, column=2, value=item.get("created_at", "")[:10])
        ws1.cell(row=row_idx, column=3, value=item.get("summary", ""))
        ws1.cell(row=row_idx, column=4, value="Open")
        ws1.cell(row=row_idx, column=5, value=proj_name)
    if action_items:
        _style_body(ws1, 2, len(action_items) + 1, len(headers1))
    ws1.column_dimensions["A"].width = 8
    ws1.column_dimensions["B"].width = 12
    ws1.column_dimensions["C"].width = 60
    ws1.column_dimensions["D"].width = 10
    ws1.column_dimensions["E"].width = 18

    # Sheet 2: Project Health
    ws2 = wb.create_sheet("Project Health")
    headers2 = ["Project", "Constraints Open", "Schedule Status", "Key Risk"]
    for col_idx, hdr in enumerate(headers2, 1):
        ws2.cell(row=1, column=col_idx, value=hdr)
    _style_header(ws2, 1, len(headers2))
    for row_idx, p in enumerate(project_health, 2):
        ws2.cell(row=row_idx, column=1, value=p["name"])
        ws2.cell(row=row_idx, column=2, value=p.get("constraints_open", 0))
        status = p.get("schedule_status", "On Track")
        cell = ws2.cell(row=row_idx, column=3, value=status)
        if status == "At Risk":
            cell.font = RED_FONT
        elif status == "Monitor":
            cell.font = AMBER_FONT
        else:
            cell.font = GREEN_FONT
        ws2.cell(row=row_idx, column=4, value=p.get("key_risk", "None identified"))
    if project_health:
        _style_body(ws2, 2, len(project_health) + 1, len(headers2))
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 18
    ws2.column_dimensions["D"].width = 40

    # Sheet 3: Constraint Movement
    ws3 = wb.create_sheet("Constraint Movement")
    headers3 = ["Project", "New", "Resolved", "Aging (>14d)", "Critical"]
    for col_idx, hdr in enumerate(headers3, 1):
        ws3.cell(row=1, column=col_idx, value=hdr)
    _style_header(ws3, 1, len(headers3))
    per_project = constraint_movement.get("per_project", {})
    row_idx = 2
    for proj_name in sorted(per_project.keys()):
        counts = per_project[proj_name]
        ws3.cell(row=row_idx, column=1, value=proj_name)
        ws3.cell(row=row_idx, column=2, value=counts.get("new", 0))
        ws3.cell(row=row_idx, column=3, value=counts.get("resolved", 0))
        ws3.cell(row=row_idx, column=4, value=counts.get("aging", 0))
        critical_cell = ws3.cell(row=row_idx, column=5, value=counts.get("critical", 0))
        if counts.get("critical", 0) > 0:
            critical_cell.font = RED_FONT
        row_idx += 1
    if per_project:
        _style_body(ws3, 2, row_idx - 1, len(headers3))
    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 10
    ws3.column_dimensions["C"].width = 12
    ws3.column_dimensions["D"].width = 14
    ws3.column_dimensions["E"].width = 12

    try:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        print(f"Excel generated: {output_path}")
        return True
    except Exception as e:
        print(f"Excel generation failed: {e}")
        return False


# ------------------------------------------------------------------
# Markdown file generation
# ------------------------------------------------------------------

def generate_morning_markdown(
    output_path: Path, date_str: str, action_items: list[dict],
    project_health: list[dict], constraint_movement: dict,
    followup_items: dict, scan_content: str | None,
) -> bool:
    """Generate morning report as Markdown file. Returns True on success."""
    now = datetime.now(CT)
    formatted_date = now.strftime("%A, %B %d, %Y")

    lines = []
    lines.append(f"# GOLIATH Morning Report")
    lines.append(f"**{formatted_date}**\n")

    # Action Items
    lines.append("---\n")
    lines.append(f"## Open Action Items ({len(action_items)} pending)\n")
    if action_items:
        by_project = {}
        for item in action_items:
            key = item.get("project_key") or "general"
            by_project.setdefault(key, []).append(item)
        general = by_project.pop("general", [])
        if general:
            lines.append("### General / System\n")
            for item in general:
                lines.append(f"- **[{item['created_at'][:10]}]** {item.get('summary', '')}")
            lines.append("")
        for proj_key in sorted(by_project.keys()):
            proj_name = PROJECTS.get(proj_key, proj_key)
            lines.append(f"### {proj_name}\n")
            for item in by_project[proj_key]:
                lines.append(f"- **[{item['created_at'][:10]}]** {item.get('summary', '')}")
            lines.append("")
    else:
        lines.append("*No open action items.*\n")

    # Project Health
    lines.append("---\n")
    lines.append(f"## Portfolio Health Summary ({len(project_health)} projects)\n")
    lines.append("| Project | POD | Schedule | Constraints | Open | Status | Key Risk |")
    lines.append("|---------|-----|----------|-------------|------|--------|----------|")
    for p in project_health:
        lines.append(
            f"| {p['name']} | {p.get('pod', 0)} | {p.get('schedule', 0)} | "
            f"{p.get('constraints', 0)} | {p.get('constraints_open', 0)} | "
            f"{p.get('schedule_status', 'On Track')} | {p.get('key_risk', 'None')[:50]} |")
    lines.append("")

    # Constraint Movement
    lines.append("---\n")
    lines.append("## Constraint Movement (24h)\n")
    total_c = constraint_movement.get("total_current", 0)
    lines.append(f"*{total_c} constraints tracked*\n")
    new_items = constraint_movement.get("new", [])
    resolved_items = constraint_movement.get("resolved", [])
    total_changes = (len(new_items) + len(resolved_items)
                     + len(constraint_movement.get("status_changed", []))
                     + len(constraint_movement.get("priority_changed", [])))
    if total_changes == 0:
        lines.append("*No changes in the last 24 hours.*\n")
    else:
        lines.append(f"**{total_changes} change(s) detected:**\n")
        if new_items:
            lines.append(f"### New ({len(new_items)})\n")
            for item in new_items[:10]:
                lines.append(f"- **[{item.get('priority', '?')}]** {item.get('project', '?')}: {item.get('description', '?')}")
            lines.append("")
        if resolved_items:
            lines.append(f"### Resolved ({len(resolved_items)})\n")
            for item in resolved_items[:10]:
                lines.append(f"- {item.get('project', '?')}: {item.get('description', '?')}")
            lines.append("")

    # Follow-Up Queue
    lines.append("---\n")
    lines.append("## Follow-Up Queue\n")
    overdue = followup_items.get("overdue", [])
    due_today = followup_items.get("due_today", [])
    if not overdue and not due_today:
        lines.append("*No follow-ups due or overdue today.*\n")
    else:
        if overdue:
            lines.append(f"### Overdue ({len(overdue)})\n")
            for item in overdue[:12]:
                lines.append(f"- **[{item.get('project_key', '?')}]** {str(item.get('commitment', ''))[:120]}")
            lines.append("")
        if due_today:
            lines.append(f"### Due Today ({len(due_today)})\n")
            for item in due_today[:12]:
                lines.append(f"- **[{item.get('project_key', '?')}]** {str(item.get('commitment', ''))[:120]}")
            lines.append("")

    # Scan
    if scan_content:
        lines.append("---\n")
        lines.append("## Latest Daily Scan\n")
        if len(scan_content) > 5000:
            lines.append(scan_content[:4800])
            lines.append("\n*... (truncated)*\n")
        else:
            lines.append(scan_content)

    lines.append("---\n")
    lines.append("*Generated by GOLIATH*")

    try:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text("\n".join(lines))
        print(f"Markdown generated: {output_path}")
        return True
    except Exception as e:
        print(f"Markdown generation failed: {e}")
        return False


# ------------------------------------------------------------------
# Telegram sending
# ------------------------------------------------------------------

def send_telegram_message(chat_id: str, text: str) -> bool:
    """Send a short text message via Telegram Bot API."""
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
    payload = {
        "chat_id": chat_id,
        "text": text,
        "disable_web_page_preview": True,
    }
    try:
        resp = requests.post(url, json=payload, timeout=30)
        if resp.status_code == 200:
            print(f"Notification sent ({len(text)} chars)")
            return True
        else:
            print(f"Failed to send notification: {resp.status_code} {resp.text}")
            return False
    except Exception as e:
        print(f"Failed to send notification: {e}")
        return False


def send_telegram_document(chat_id: str, file_path: Path, caption: str = "") -> bool:
    """Send a file as a Telegram document attachment."""
    if not file_path.exists():
        print(f"Cannot send document -- file not found: {file_path}")
        return False

    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendDocument"
    try:
        with open(file_path, "rb") as f:
            data = {"chat_id": chat_id}
            if caption:
                data["caption"] = caption[:1024]
                data["parse_mode"] = "HTML"
            files = {"document": (file_path.name, f)}
            resp = requests.post(url, data=data, files=files, timeout=60)

        if resp.status_code == 200:
            print(f"Sent document: {file_path.name}")
            return True
        else:
            print(f"Failed to send {file_path.name}: {resp.status_code} {resp.text[:200]}")
            return False
    except Exception as e:
        print(f"Failed to send {file_path.name}: {e}")
        return False


# ------------------------------------------------------------------
# Main
# ------------------------------------------------------------------

def main():
    if not TELEGRAM_BOT_TOKEN:
        print("Error: TELEGRAM_BOT_TOKEN not set in .env")
        return 1

    if not TELEGRAM_CHAT_ID:
        print("Error: REPORT_CHAT_ID (or ALLOWED_CHAT_IDS) not set in .env")
        return 1

    chat_id = TELEGRAM_CHAT_ID.split(",")[0].strip()
    now = datetime.now(CT)
    date_iso = now.strftime("%Y-%m-%d")
    date_display = now.strftime("%A, %B %d, %Y")

    print(f"[{now.isoformat()}] Gathering morning report data...")

    # ---- Gather structured data ----
    action_items = gather_open_action_items()
    project_health = gather_project_health()
    constraint_movement = gather_constraint_movement()
    followup_items = gather_followup_items()
    scan_content = get_latest_scan_content()

    # ---- Build short notification ----
    n_items = len(action_items)
    n_changes = len(constraint_movement.get("new", [])) + len(constraint_movement.get("resolved", []))
    n_overdue = len(followup_items.get("overdue", []))
    n_due = len(followup_items.get("due_today", []))

    summary_parts = []
    if n_items > 0:
        summary_parts.append(f"{n_items} open action items")
    if n_changes > 0:
        summary_parts.append(f"{n_changes} constraint changes overnight")
    if n_overdue > 0:
        summary_parts.append(f"{n_overdue} overdue follow-ups")
    elif n_due > 0:
        summary_parts.append(f"{n_due} follow-ups due today")
    if not summary_parts:
        summary_parts.append("all clear across the portfolio")

    notification = (
        f"\u2600\ufe0f Morning report ready \u2014 {date_display}.\n"
        f"{'; '.join(summary_parts)}.\n"
        f"Files attached below."
    )

    # ---- Generate files ----
    report_dir = REPORTS_DIR
    report_dir.mkdir(parents=True, exist_ok=True)

    pdf_path = report_dir / f"{date_iso}-morning-report.pdf"
    xlsx_path = report_dir / f"{date_iso}-morning-report.xlsx"
    md_path = report_dir / f"{date_iso}-morning-report.md"

    pdf_ok = generate_morning_pdf(
        pdf_path, date_iso, action_items, project_health,
        constraint_movement, followup_items, scan_content,
    )
    xlsx_ok = generate_morning_excel(
        xlsx_path, date_iso, action_items, project_health, constraint_movement,
    )
    md_ok = generate_morning_markdown(
        md_path, date_iso, action_items, project_health,
        constraint_movement, followup_items, scan_content,
    )

    # ---- Send notification ----
    send_telegram_message(chat_id, notification)

    # ---- Send file attachments ----
    files_sent = 0
    if pdf_ok and pdf_path.exists():
        if send_telegram_document(chat_id, pdf_path,
                                  caption=f"<b>Morning Report (PDF)</b> \u2014 {date_iso}"):
            files_sent += 1
    if xlsx_ok and xlsx_path.exists():
        if send_telegram_document(chat_id, xlsx_path,
                                  caption=f"<b>Morning Report (Excel)</b> \u2014 {date_iso}"):
            files_sent += 1
    if md_ok and md_path.exists():
        if send_telegram_document(chat_id, md_path,
                                  caption=f"<b>Morning Report (Markdown)</b> \u2014 {date_iso}"):
            files_sent += 1

    print(f"Morning report complete: {files_sent}/3 files sent.")
    return 0 if files_sent > 0 else 1


if __name__ == "__main__":
    sys.exit(main())
