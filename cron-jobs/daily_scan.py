#!/usr/bin/env python3
"""
Daily Project Scan — 11 PM CT (run by internal scheduler)

Scans POD, Schedule, and Constraints for all 12 projects.
Extracts text from PDFs and Excel files so Claude can analyze real data.
Generates a report with findings and questions for the site team.
Saves report to cron-jobs/reports/YYYY-MM-DD_daily_scan.md
"""

import os
import subprocess
import sys
from datetime import datetime
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent
PROJECTS_DIR = REPO_ROOT / "projects"
REPORTS_DIR = REPO_ROOT / "cron-jobs" / "reports"
SCAN_FOLDERS = ["pod", "schedule", "constraints"]

PROJECTS = [
    "union-ridge", "duff", "salt-branch", "blackford",
    "delta-bobcat", "tehuacana", "three-rivers", "scioto-ridge",
    "mayes", "graceland", "pecan-prairie", "duffy-bess",
]

# File size limits for extraction (bytes)
MAX_PDF_SIZE = 10_000_000   # 10 MB
MAX_EXCEL_SIZE = 10_000_000  # 10 MB
MAX_TEXT_SIZE = 500_000      # 500 KB
MAX_EXTRACTED_CHARS = 15_000  # Max chars per file to include in prompt


def extract_pdf_text(filepath: Path) -> str | None:
    """Extract text content from a PDF file using pdfminer.six."""
    if filepath.stat().st_size > MAX_PDF_SIZE:
        return None
    try:
        from pdfminer.high_level import extract_text
        text = extract_text(str(filepath))
        if text and text.strip():
            return text.strip()[:MAX_EXTRACTED_CHARS]
        return None
    except Exception as e:
        return f"[PDF extraction failed: {e}]"


def extract_excel_text(filepath: Path) -> str | None:
    """Extract text content from an Excel file using openpyxl."""
    if filepath.stat().st_size > MAX_EXCEL_SIZE:
        return None
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
        all_text = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            all_text.append(f"--- Sheet: {sheet_name} ---")
            row_count = 0
            for row in ws.iter_rows(values_only=True):
                if row_count > 500:  # Cap rows per sheet
                    all_text.append(f"... ({row_count}+ rows, truncated)")
                    break
                # Convert row to tab-separated string, skip empty rows
                values = [str(v) if v is not None else "" for v in row]
                if any(v.strip() for v in values):
                    all_text.append("\t".join(values))
                    row_count += 1
        wb.close()
        result = "\n".join(all_text)
        if result.strip():
            return result[:MAX_EXTRACTED_CHARS]
        return None
    except Exception as e:
        return f"[Excel extraction failed: {e}]"


def extract_xer_text(filepath: Path) -> str | None:
    """Extract text from XER files (P6 schedule exports — plain text format)."""
    if filepath.stat().st_size > MAX_TEXT_SIZE:
        return None
    try:
        text = filepath.read_text(errors="replace")
        if text.strip():
            return text.strip()[:MAX_EXTRACTED_CHARS]
        return None
    except Exception as e:
        return f"[XER extraction failed: {e}]"


def extract_file_content(filepath: Path) -> tuple[str | None, str]:
    """
    Attempt to extract readable content from a file.
    Returns (content_or_none, file_type_label).
    """
    suffix = filepath.suffix.lower()

    if suffix == ".pdf":
        return extract_pdf_text(filepath), "PDF"

    elif suffix in (".xlsx", ".xlsm", ".xls"):
        return extract_excel_text(filepath), "Excel"

    elif suffix == ".xer":
        return extract_xer_text(filepath), "XER (P6 Schedule)"

    elif suffix in (".md", ".txt", ".csv", ".json", ".xml", ".log"):
        if filepath.stat().st_size < MAX_TEXT_SIZE:
            try:
                content = filepath.read_text(errors="replace")
                if content.strip():
                    return content.strip()[:MAX_EXTRACTED_CHARS], "Text"
            except Exception:
                pass
        return None, "Text"

    else:
        return None, f"Binary ({suffix})"


def collect_project_data() -> str:
    """Walk each project's POD, Schedule, Constraints folders and extract content."""
    sections = []

    for project_key in PROJECTS:
        project_path = PROJECTS_DIR / project_key
        project_lines = [f"## {project_key}"]
        has_readable_data = False

        for folder in SCAN_FOLDERS:
            folder_path = project_path / folder
            if not folder_path.exists():
                project_lines.append(f"### {folder}/\n_Folder missing_\n")
                continue

            files = [
                f for f in sorted(folder_path.rglob("*"))
                if f.is_file() and f.name != ".gitkeep"
            ]

            if not files:
                project_lines.append(f"### {folder}/\n_No data files_\n")
                continue

            project_lines.append(f"### {folder}/ ({len(files)} file(s))")

            for fp in files:
                rel = fp.relative_to(project_path)
                size_kb = fp.stat().st_size / 1024
                content, file_type = extract_file_content(fp)

                project_lines.append(f"- `{rel}` ({size_kb:.1f} KB, {file_type})")

                if content and not content.startswith("["):
                    has_readable_data = True
                    # Include extracted content
                    project_lines.append(
                        f"<extracted_content file=\"{rel}\">\n{content}\n</extracted_content>"
                    )
                elif content and content.startswith("["):
                    # Extraction error message
                    project_lines.append(f"  {content}")
                else:
                    project_lines.append(
                        f"  _Could not extract text content from this {file_type} file._"
                    )

            project_lines.append("")

        if not has_readable_data:
            project_lines.append(
                "\n⚠️ NO READABLE DATA EXTRACTED for this project. "
                "Files exist but no text could be extracted from them. "
                "DO NOT fabricate analysis — report 'no readable data available' for this project."
            )

        sections.append("\n".join(project_lines))

    return "\n---\n".join(sections)


def run_claude_scan(project_data: str) -> str:
    """Send collected data to Claude CLI for analysis and report generation."""
    today_str = datetime.now().strftime('%B %d, %Y')
    prompt = f"""You are the GOLIATH DSC Operations Agent running a scheduled end-of-day scan.

Below is the current state of POD, Schedule, and Constraints data across all 12 projects.
Data has been extracted from PDF, Excel, XER, and text files where possible.

## CRITICAL RULES — READ THESE FIRST

1. **DO NOT HALLUCINATE OR FABRICATE DATA.** If a project has no extracted content (only file names
   and sizes), you MUST report: "No readable data available — files exist but could not be extracted.
   Manual review needed." Do NOT invent numbers, dates, forecasts, or analysis.

2. **Only analyze what you can actually see.** If you have extracted text from a PDF or Excel file,
   analyze that data. If you only have a filename and file size, say so and move on.

3. **Be explicit about data sources.** For every finding, cite the specific file and data point
   you're referencing. If you can't cite a source, don't make the claim.

4. **Distinguish between data and inference.** When you calculate or infer something from the data,
   label it clearly: "Based on the extracted data from [file], ..."

5. **When in doubt, say "insufficient data."** It is far better to say "I don't have enough data
   to assess this" than to generate a plausible-sounding but fabricated analysis.

## Report Format

# Daily Scan Report — {today_str}

## Executive Summary
(2-3 sentences on overall portfolio health BASED ON AVAILABLE DATA. If most projects lack
readable data, say that clearly.)

## Findings by Project
For EACH project:
- If readable data was extracted: Analyze POD, Schedule, and Constraints based on the actual data
- If NO readable data: Report "No readable data available — [X] files exist but content could
  not be extracted. File types: [list types]. Recommend uploading text/CSV versions or having
  subagents analyze these files interactively."
- NEVER fabricate metrics, dates, or projections

## Portfolio-Level Concerns
(Cross-project risks based on ACTUAL DATA ONLY. If insufficient data, say so.)

## Questions for Site Teams
For projects WITH data: 2-4 specific, data-driven questions
For projects WITHOUT data: "What format are your reports in? Can you provide CSV/text exports?"

## Action Items
(Recommended next steps based on what was actually found)

## Data Coverage Summary
List how many projects had readable data vs. how many had only filenames.
This helps track our data ingestion progress.

---

PROJECT DATA:
{project_data}
"""

    env = os.environ.copy()
    env.pop("CLAUDECODE", None)

    cmd = [
        "claude",
        "--print",
        "--dangerously-skip-permissions",
        "--output-format", "text",
        prompt,
    ]

    result = subprocess.run(
        cmd,
        capture_output=True,
        text=True,
        timeout=600,
        cwd=str(REPO_ROOT),
        env=env,
    )

    if result.returncode != 0:
        return f"# Scan Failed\n\n```\n{result.stderr[:2000]}\n```"

    return result.stdout.strip() if result.stdout.strip() else "# Scan returned empty response"


def main():
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    today = datetime.now().strftime("%Y-%m-%d")
    report_path = REPORTS_DIR / f"{today}_daily_scan.md"

    print(f"[{datetime.now().isoformat()}] Starting daily scan...")

    # Collect data from all projects (with PDF/Excel extraction)
    print("Collecting project data (extracting from PDFs, Excel, XER, text files)...")
    project_data = collect_project_data()
    print(f"Collected {len(project_data)} chars of project data")

    # Run Claude analysis
    print("Running Claude analysis...")
    report = run_claude_scan(project_data)

    # Save report
    report_path.write_text(report)
    print(f"Report saved: {report_path}")
    print(f"Report size: {len(report)} chars")

    return 0


if __name__ == "__main__":
    sys.exit(main())
