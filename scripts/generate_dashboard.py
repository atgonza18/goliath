#!/usr/bin/env python3
"""
Generate the Goliath Portfolio Dashboard Excel report.
Output: /opt/goliath/reports/2026-02-26-portfolio-dashboard.xlsx
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_PATH = "/opt/goliath/reports/2026-02-26-portfolio-dashboard.xlsx"
TITLE_TEXT = "GOLIATH Portfolio Report \u2014 February 26, 2026"

# ── Color palette ──────────────────────────────────────────────────────────
DARK_BLUE_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
GRAY_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
TITLE_FILL = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")

WHITE_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(color="FFFFFF", bold=True, size=14)
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
BODY_FONT = Font(size=10)
BOLD_BODY_FONT = Font(size=10, bold=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def add_title_row(ws, num_cols):
    """Merge row 1 across all columns and write the report title."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    cell = ws.cell(row=1, column=1, value=TITLE_TEXT)
    cell.font = TITLE_FONT
    cell.fill = TITLE_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32
    # Fill merged area styling
    for c in range(2, num_cols + 1):
        ws.cell(row=1, column=c).fill = TITLE_FILL


def write_headers(ws, headers, row=2):
    """Write header row with dark-blue background and white bold text."""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = DARK_BLUE_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 22


def write_data_row(ws, row_idx, values, fill=None, bold_col=None):
    """Write a single data row, optionally applying a fill and bolding a column."""
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = BODY_FONT
        cell.alignment = CENTER if col_idx > 1 else LEFT
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill
        if bold_col and col_idx == bold_col:
            cell.font = BOLD_BODY_FONT


def auto_col_widths(ws, headers, min_width=15, max_width=35):
    """Set column widths based on header length with min/max bounds."""
    for col_idx, header in enumerate(headers, 1):
        # Scan data for widest cell
        widest = len(str(header))
        for row in ws.iter_rows(min_row=3, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    widest = max(widest, len(str(cell.value)))
        width = max(min_width, min(widest + 4, max_width))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ══════════════════════════════════════════════════════════════════════════
#  SHEET 1 — Portfolio Summary
# ══════════════════════════════════════════════════════════════════════════

def build_portfolio_summary(wb):
    ws = wb.active
    ws.title = "Portfolio Summary"

    headers = ["Project", "MW DC", "Contractor", "Overall Float",
               "COD/SC Target", "Risk Level", "Phase"]
    num_cols = len(headers)

    add_title_row(ws, num_cols)
    write_headers(ws, headers, row=2)

    data = [
        ("Blackford Solar", 211.21, "White/MasTec", -53, "Aug 6, 2026", "HIGH", "Active Construction"),
        ("Duff Solar", 138, "White/MasTec", -158, "Jul 21, 2026", "EXTREME", "Active Construction"),
        ("Pecan Prairie North", 407, "Wanzek", -84, "Sep 9, 2027", "MODERATE", "Early Construction"),
        ("Pecan Prairie South", 188, "Wanzek", -10, "May 10, 2027", "LOW-MOD", "Early Construction"),
        ("Scioto Ridge", "\u2014", "\u2014", "\u2014", "\u2014", "NO DATA", "Data Error (Blackford dupes)"),
        ("Union Ridge", "\u2014", "\u2014", "\u2014", "\u2014", "NO DATA", "No Schedule Uploaded"),
        ("Salt Branch", "\u2014", "\u2014", "\u2014", "\u2014", "NO DATA", "No Schedule Uploaded"),
        ("Delta Bobcat", "\u2014", "\u2014", "\u2014", "\u2014", "NO DATA", "No Schedule Uploaded"),
        ("Tehuacana", "\u2014", "\u2014", "\u2014", "\u2014", "NO DATA", "No Schedule Uploaded"),
        ("Three Rivers", "\u2014", "\u2014", "\u2014", "\u2014", "NO DATA", "No Schedule Uploaded"),
        ("Mayes", "\u2014", "\u2014", "\u2014", "\u2014", "NO DATA", "No Schedule Uploaded"),
        ("Graceland", "\u2014", "\u2014", "\u2014", "\u2014", "NO DATA", "No Schedule Uploaded"),
        ("Duffy BESS", "\u2014", "\u2014", "\u2014", "\u2014", "NO DATA", "No Schedule Uploaded"),
    ]

    risk_fills = {
        "EXTREME": RED_FILL,
        "HIGH": RED_FILL,
        "MODERATE": YELLOW_FILL,
        "LOW-MOD": GREEN_FILL,
        "NO DATA": GRAY_FILL,
    }

    for i, row_data in enumerate(data):
        row_idx = i + 3  # data starts at row 3
        risk = row_data[5]
        fill = risk_fills.get(risk, None)
        write_data_row(ws, row_idx, row_data, fill=fill, bold_col=1)

    auto_col_widths(ws, headers)
    ws.sheet_properties.tabColor = "1F3864"


# ══════════════════════════════════════════════════════════════════════════
#  SHEET 2 — Blackford Milestones
# ══════════════════════════════════════════════════════════════════════════

def build_blackford(wb):
    ws = wb.create_sheet("Blackford Milestones")

    headers = ["Milestone", "Contract Date", "Forecast Date", "Float (Days)", "Status"]
    num_cols = len(headers)

    add_title_row(ws, num_cols)
    write_headers(ws, headers, row=2)

    data = [
        ("Backfeed/Interconnect", "Jan 19, 2026", "Jun 1, 2026", -94, "CRITICAL"),
        ("HV Works Completion", "Jan 19, 2026", "Jun 1, 2026", -94, "CRITICAL"),
        ("1st Circuit MC", "Nov 20, 2025", "Mar 25, 2026", -79, "CRITICAL"),
        ("2nd Circuit MC", "Dec 18, 2025", "Apr 13, 2026", -17, "Behind"),
        ("3rd Circuit MC", "Jan 20, 2026", "Apr 27, 2026", -25, "Behind"),
        ("4th Circuit MC", "Feb 10, 2026", "May 7, 2026", -31, "Behind"),
        ("5th Circuit MC", "Mar 5, 2026", "May 21, 2026", -38.5, "Behind"),
        ("6th Circuit MC", "Mar 26, 2026", "Jun 2, 2026", -44, "Behind"),
        ("7th Circuit MC", "Apr 16, 2026", "Jun 10, 2026", -48, "Behind"),
        ("8th Circuit MC", "Apr 29, 2026", "Jun 18, 2026", -52, "Behind"),
        ("Project MC", "\u2014", "Jun 19, 2026", -20, "Behind"),
        ("COD", "Jun 5, 2026", "Aug 6, 2026", -53, "CRITICAL"),
        ("Substantial Completion", "Jun 5, 2026", "Aug 6, 2026", -53, "CRITICAL"),
        ("Final Completion", "Jul 6, 2026", "Sep 18, 2026", -53, "CRITICAL"),
    ]

    status_fills = {
        "CRITICAL": RED_FILL,
        "Behind": YELLOW_FILL,
    }

    for i, row_data in enumerate(data):
        row_idx = i + 3
        status = row_data[4]
        fill = status_fills.get(status, None)
        write_data_row(ws, row_idx, row_data, fill=fill, bold_col=1)

    auto_col_widths(ws, headers)
    ws.sheet_properties.tabColor = "C00000"


# ══════════════════════════════════════════════════════════════════════════
#  SHEET 3 — Duff Milestones
# ══════════════════════════════════════════════════════════════════════════

def build_duff(wb):
    ws = wb.create_sheet("Duff Milestones")

    headers = ["Milestone", "LD Date", "Forecast Date", "Float (Days)", "Status"]
    num_cols = len(headers)

    add_title_row(ws, num_cols)
    write_headers(ws, headers, row=2)

    data = [
        ("HV Completion", "Oct 30, 2025", "Jan 8, 2026", "\u2014", "COMPLETE"),
        ("1st Circuit MC", "Oct 23, 2025", "Apr 14, 2026", -117, "CRITICAL"),
        ("2nd Circuit MC", "\u2014", "May 27, 2026", -147, "CRITICAL"),
        ("3rd Circuit MC", "Nov 5, 2025", "Jun 24, 2026", -158, "CRITICAL"),
        ("1st Circuit Energization", "\u2014", "May 29, 2026", -116, "CRITICAL"),
        ("2nd Circuit Energization", "\u2014", "Jun 19, 2026", -126, "CRITICAL"),
        ("3rd Circuit Energization", "\u2014", "Jun 26, 2026", -125, "CRITICAL"),
        ("1st Circuit Turnover", "Jan 22, 2026", "Jun 25, 2026", -117, "CRITICAL"),
        ("2nd Circuit Turnover", "\u2014", "Jul 13, 2026", -118, "CRITICAL"),
        ("3rd Circuit Turnover", "Feb 2, 2026", "Jul 20, 2026", -118, "CRITICAL"),
        ("Substantial Completion", "Feb 6, 2026", "Jul 21, 2026", -115, "CRITICAL"),
        ("Final Completion", "\u2014", "Aug 18, 2026", 0, "On Track"),
    ]

    status_fills = {
        "CRITICAL": RED_FILL,
        "COMPLETE": GREEN_FILL,
        "On Track": GREEN_FILL,
    }

    for i, row_data in enumerate(data):
        row_idx = i + 3
        status = row_data[4]
        fill = status_fills.get(status, None)
        write_data_row(ws, row_idx, row_data, fill=fill, bold_col=1)

    auto_col_widths(ws, headers)
    ws.sheet_properties.tabColor = "C00000"


# ══════════════════════════════════════════════════════════════════════════
#  SHEET 4 — Pecan Prairie Milestones
# ══════════════════════════════════════════════════════════════════════════

def build_pecan(wb):
    ws = wb.create_sheet("Pecan Prairie Milestones")

    headers = ["Site", "Milestone", "Contract Date", "Forecast Date", "Status"]
    num_cols = len(headers)

    add_title_row(ws, num_cols)
    write_headers(ws, headers, row=2)

    data = [
        ("North", "Mobilization", "\u2014", "Sep 29, 2025", "COMPLETE"),
        ("North", "Backfeed Available", "\u2014", "Oct 30, 2026", "Future"),
        ("North", "Substation Energization", "Nov 24, 2026", "Nov 25, 2026", "On Track"),
        ("North", "MC 250 MW AC", "Dec 31, 2026", "Mar 5, 2027", "Watch (-63d)"),
        ("North", "Full MC", "Feb 19, 2027", "May 13, 2027", "Watch (-83d)"),
        ("North", "Substantial Completion", "Apr 30, 2027", "Jul 23, 2027", "Watch (-84d)"),
        ("North", "COD/Final Acceptance", "Jun 17, 2027", "Sep 9, 2027", "Watch (-84d)"),
        ("South", "Mobilization", "Sep 29, 2025", "Mar 2, 2026", "Delayed"),
        ("South", "T-Line Energization", "Nov 17, 2026", "Jan 7, 2027", "Watch"),
        ("South", "MC", "Mar 12, 2027", "Mar 22, 2027", "On Track"),
        ("South", "Substantial Completion", "Apr 30, 2027", "May 10, 2027", "On Track"),
        ("South", "Final Completion", "May 28, 2027", "Jun 30, 2027", "On Track"),
    ]

    status_fills = {
        "COMPLETE": GREEN_FILL,
        "On Track": GREEN_FILL,
        "Future": None,
        "Watch": YELLOW_FILL,
        "Watch (-63d)": YELLOW_FILL,
        "Watch (-83d)": YELLOW_FILL,
        "Watch (-84d)": YELLOW_FILL,
        "Delayed": RED_FILL,
    }

    for i, row_data in enumerate(data):
        row_idx = i + 3
        status = row_data[4]
        fill = status_fills.get(status, None)
        write_data_row(ws, row_idx, row_data, fill=fill, bold_col=2)

    auto_col_widths(ws, headers)
    ws.sheet_properties.tabColor = "2E75B6"


# ══════════════════════════════════════════════════════════════════════════
#  Main
# ══════════════════════════════════════════════════════════════════════════

def main():
    wb = openpyxl.Workbook()

    build_portfolio_summary(wb)
    build_blackford(wb)
    build_duff(wb)
    build_pecan(wb)

    # Freeze panes: freeze below headers (row 2) on every sheet
    for ws in wb.worksheets:
        ws.freeze_panes = "A3"

    wb.save(OUTPUT_PATH)
    print(f"Dashboard saved to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
